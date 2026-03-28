import base64
import binascii
import csv
import difflib
import io
import logging
import os
import unicodedata
from datetime import date, datetime, time

from odoo import _, fields, models
from odoo.exceptions import UserError
from odoo.fields import Command

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - depends on server runtime
    load_workbook = None

_logger = logging.getLogger(__name__)


class WgsSubscriptionImportWizard(models.TransientModel):
    _name = 'wgs.subscription.import.wizard'
    _description = 'Importador de suscripciones vigentes'

    file_data = fields.Binary(string='Archivo', required=True)
    file_name = fields.Char(string='Nombre de archivo')
    batch_name = fields.Char(
        string='Lote',
        default=lambda self: fields.Datetime.now().strftime('WGS Import %Y-%m-%d %H:%M:%S'),
        required=True,
        help='Etiqueta para identificar las órdenes generadas por esta corrida.',
    )
    dry_run = fields.Boolean(
        string='Solo analizar (dry-run)',
        default=True,
        help='Si está activado, valida filas y muestra el resultado sin crear ni actualizar órdenes.',
    )
    update_existing = fields.Boolean(
        string='Actualizar suscripciones ya importadas',
        default=True,
        help='Reutiliza la llave de importación para actualizar una orden existente en vez de duplicarla.',
    )
    skip_non_current = fields.Boolean(
        string='Omitir filas no vigentes hoy',
        default=True,
        help='Solo procesa filas con inicio menor o igual a hoy y fin mayor o igual a hoy.',
    )
    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        required=True,
        default=lambda self: self.env.company,
    )
    active_state_mode = fields.Selection(
        [
            ('progress', 'En progreso'),
            ('renew', 'Por renovar'),
        ],
        string='Estado activo a aplicar',
        default='progress',
        required=True,
    )
    result_summary = fields.Char(string='Resumen', readonly=True)
    result_log = fields.Text(string='Resultado', readonly=True)
    _HEADER_ALIASES = {
        'partner_id': ('partner_id', 'id_partner'),
        'customer_id': ('id_cliente', 'cliente_id', 'customer_id', 'client_id', 'codigo_cliente_propietario'),
        'ref': ('ref', 'codigo', 'codigo_cliente', 'codigo_usuario', 'partner_ref'),
        'xml_id': ('xml_id', 'external_id', 'id_externo', 'record_id'),
        'email': ('email', 'correo', 'correo_electronico', 'mail'),
        'mobile': ('mobile', 'movil', 'celular', 'telefono_movil'),
        'phone': ('phone', 'telefono', 'telefono_fijo'),
        'vat': ('vat', 'rfc', 'documento', 'identificacion'),
        'name': ('name', 'nombre', 'cliente', 'usuario', 'partner'),
        'plan': ('plan', 'paquete', 'producto', 'suscripcion', 'subscription', 'membership'),
        'subscription_plan': (
            'subscription_plan',
            'billing_plan',
            'plan_facturacion',
            'plan_cobro',
            'periodicidad',
        ),
        'start_date': (
            'start_date',
            'inicio',
            'fecha_inicio',
            'inicio_vigencia',
            'vigencia_inicio',
        ),
        'end_date': (
            'end_date',
            'fin',
            'fecha_fin',
            'fin_vigencia',
            'vigencia_fin',
        ),
        'state': ('state', 'estado', 'subscription_state'),
        'price': ('price', 'precio', 'importe', 'monto'),
        'quantity': ('quantity', 'qty', 'cantidad'),
        'participants': ('participants', 'participantes', 'participantes', 'nombres_participantes'),
    }

    def action_process_file(self):
        self.ensure_one()
        rows = self._load_rows_from_upload()
        if not rows:
            raise UserError(_('El archivo no contiene filas de datos.'))

        today = fields.Date.context_today(self)
        active_state_value = self._resolve_subscription_state_value(self.active_state_mode)
        if not active_state_value:
            raise UserError(_('No se pudo resolver un valor válido para subscription_state en este entorno.'))

        partner_cache = {}
        product_cache = {}
        subscription_plan_cache = {}
        order_model = self.env['sale.order'].sudo().with_company(self.company_id)

        counters = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }
        log_lines = []

        for row_number, row_data in rows:
            try:
                with self.env.cr.savepoint():
                    normalized = self._normalize_row(row_data)
                    if not normalized:
                        counters['skipped'] += 1
                        log_lines.append(_('Fila %(row)s: omitida porque está vacía.') % {'row': row_number})
                        continue

                    start_date = self._parse_date_value(normalized.get('start_date'), field_label='inicio', row_number=row_number)
                    end_date = self._parse_date_value(normalized.get('end_date'), field_label='fin', row_number=row_number)
                    if not start_date or not end_date:
                        raise UserError(
                            _(
                                'Fila %(row)s: las columnas de inicio y fin son obligatorias.'
                            )
                            % {'row': row_number}
                        )
                    if end_date < start_date:
                        raise UserError(
                            _(
                                'Fila %(row)s: la fecha de fin %(end)s no puede ser menor que la de inicio %(start)s.'
                            )
                            % {
                                'row': row_number,
                                'start': fields.Date.to_string(start_date),
                                'end': fields.Date.to_string(end_date),
                            }
                        )

                    if self.skip_non_current and (start_date > today or end_date < today):
                        counters['skipped'] += 1
                        log_lines.append(
                            _(
                                'Fila %(row)s: omitida por no estar vigente hoy (%(start)s -> %(end)s).'
                            )
                            % {
                                'row': row_number,
                                'start': fields.Date.to_string(start_date),
                                'end': fields.Date.to_string(end_date),
                            }
                        )
                        continue

                    state_value = self._resolve_import_subscription_state_value(
                        start_date=start_date,
                        end_date=end_date,
                        today=today,
                        active_state_value=active_state_value,
                    )
                    partner = self._resolve_partner(normalized, partner_cache, row_number=row_number)
                    product = self._resolve_subscription_product(normalized.get('plan'), product_cache, row_number=row_number)
                    subscription_plan = self._resolve_subscription_plan(
                        normalized.get('subscription_plan'),
                        subscription_plan_cache,
                        row_number=row_number,
                    )
                    pricing_context = self._resolve_recurring_context(
                        product=product,
                        subscription_plan=subscription_plan,
                        start_date=start_date,
                        fallback_price=normalized.get('price'),
                    )
                    subscription_plan = pricing_context.get('plan') or subscription_plan
                    self._ensure_subscription_plan_ready(
                        product=product,
                        subscription_plan=subscription_plan,
                        row_number=row_number,
                    )
                    recurring_pricing_id = pricing_context.get('pricing_id') or False
                    next_billing_date = pricing_context.get('next_billing_date') or False
                    participant_ids = self._resolve_participants(
                        raw_value=normalized.get('participants'),
                        owner=partner,
                        row_number=row_number,
                    )
                    quantity = self._parse_quantity_value(normalized.get('quantity'))
                    price_unit = self._parse_price_value(
                        normalized.get('price'),
                        fallback=float(pricing_context.get('price') or product.list_price or 0.0),
                    )
                    source_key = self._build_source_key(partner, product, start_date)

                    existing_order = order_model.search(
                        [
                            ('wgs_import_source_key', '=', source_key),
                            ('state', '!=', 'cancel'),
                        ],
                        limit=1,
                        order='id desc',
                    )
                    action_label = 'preview'

                    if existing_order and not self.update_existing:
                        counters['skipped'] += 1
                        log_lines.append(
                            _(
                                'Fila %(row)s: ya existe %(order)s para %(partner)s y update_existing está desactivado.'
                            )
                            % {
                                'row': row_number,
                                'order': existing_order.display_name,
                                'partner': partner.display_name,
                            }
                        )
                        continue

                    if not self.dry_run:
                        if existing_order:
                            self._update_existing_subscription_order(
                                order=existing_order,
                                partner=partner,
                                product=product,
                                subscription_plan=subscription_plan,
                                start_date=start_date,
                                end_date=end_date,
                                state_value=state_value,
                                price_unit=price_unit,
                                quantity=quantity,
                                participant_ids=participant_ids,
                                recurring_pricing_id=recurring_pricing_id,
                                next_billing_date=next_billing_date,
                                source_key=source_key,
                            )
                            counters['updated'] += 1
                            action_label = 'updated'
                        else:
                            order = self._create_subscription_order(
                                partner=partner,
                                product=product,
                                subscription_plan=subscription_plan,
                                start_date=start_date,
                                end_date=end_date,
                                state_value=state_value,
                                price_unit=price_unit,
                                quantity=quantity,
                                participant_ids=participant_ids,
                                recurring_pricing_id=recurring_pricing_id,
                                next_billing_date=next_billing_date,
                                source_key=source_key,
                            )
                            counters['created'] += 1
                            action_label = order.display_name
                    else:
                        action_label = existing_order.display_name if existing_order else _('se crearía')
                        if existing_order:
                            counters['updated'] += 1
                        else:
                            counters['created'] += 1

                    log_lines.append(
                        _(
                        'Fila %(row)s: %(partner)s -> %(product)s (%(start)s -> %(end)s) [%(action)s].'
                        )
                        % {
                            'row': row_number,
                            'partner': partner.display_name,
                            'product': product.display_name,
                            'start': fields.Date.to_string(start_date),
                            'end': fields.Date.to_string(end_date),
                            'action': action_label,
                        }
                    )
            except Exception as error:  # pragma: no cover - integration behavior
                counters['errors'] += 1
                log_lines.append(_('Fila %(row)s: ERROR: %(error)s') % {'row': row_number, 'error': str(error)})
                _logger.warning('WGS import row failed row=%s error=%s', row_number, error, exc_info=True)

        mode_label = _('Dry-run') if self.dry_run else _('Importación real')
        self.result_summary = _(
            '%(mode)s: creadas=%(created)s actualizadas=%(updated)s omitidas=%(skipped)s errores=%(errors)s'
        ) % {
            'mode': mode_label,
            'created': counters['created'],
            'updated': counters['updated'],
            'skipped': counters['skipped'],
            'errors': counters['errors'],
        }
        self.result_log = '\n'.join(log_lines[:300])

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def _load_rows_from_upload(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('Debes subir un archivo para continuar.'))

        try:
            raw = base64.b64decode(self.file_data, validate=True)
        except (binascii.Error, ValueError) as error:
            raise UserError(_('No se pudo leer el archivo cargado: %s') % error) from error

        file_name = (self.file_name or '').strip().lower()
        if file_name.endswith('.csv'):
            return self._load_csv_rows(raw)
        if file_name.endswith('.xlsx'):
            return self._load_xlsx_rows(raw)
        if raw[:2] == b'PK':
            return self._load_xlsx_rows(raw)
        return self._load_csv_rows(raw)

    def _load_csv_rows(self, raw):
        text = False
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is False:
            raise UserError(_('No se pudo decodificar el CSV. Usa UTF-8 o Latin-1.'))

        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ','

        reader = csv.reader(io.StringIO(text), dialect=dialect)
        return self._rows_from_iterable(reader)

    def _load_xlsx_rows(self, raw):
        if not load_workbook:
            raise UserError(_('El runtime de Odoo no tiene openpyxl instalado para leer archivos .xlsx.'))
        workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        worksheet = workbook.active
        return self._rows_from_iterable(worksheet.iter_rows(values_only=True))

    def _rows_from_iterable(self, iterable):
        header = False
        rows = []
        for idx, row in enumerate(iterable, start=1):
            values = list(row or [])
            if not header:
                header = [self._normalize_token(value) for value in values]
                continue
            if not any(not self._is_empty_cell(value) for value in values):
                continue
            row_dict = {}
            for position, value in enumerate(values):
                key = header[position] if position < len(header) else 'column_%s' % position
                row_dict[key] = value
            rows.append((idx, row_dict))
        return rows

    def _normalize_row(self, row_data):
        row = {}
        for canonical_key, aliases in self._HEADER_ALIASES.items():
            for alias in aliases:
                value = row_data.get(self._normalize_token(alias))
                if self._is_empty_cell(value):
                    continue
                row[canonical_key] = value
                break
        return row

    def _resolve_partner(self, row, cache, row_number):
        raw_candidates = [
            ('partner_id', row.get('partner_id')),
            ('customer_id', row.get('customer_id')),
            ('xml_id', row.get('xml_id')),
            ('ref', row.get('ref')),
            ('email', row.get('email')),
            ('mobile', row.get('mobile')),
            ('phone', row.get('phone')),
            ('vat', row.get('vat')),
            ('name', row.get('name')),
        ]
        cache_key = tuple((key, self._cacheable_value(value)) for key, value in raw_candidates if not self._is_empty_cell(value))
        if cache_key in cache:
            return cache[cache_key]

        partner = False
        for key, value in raw_candidates:
            if self._is_empty_cell(value):
                continue
            partner = self._partner_from_candidate(key, value)
            if partner:
                break

        if not partner:
            raise UserError(
                _(
                    'Fila %(row)s: no pude identificar al usuario. Usa partner_id, ID Cliente, external_id/xml_id, ref, email, teléfono o nombre.'
                )
                % {'row': row_number}
            )

        cache[cache_key] = partner
        return partner

    def _partner_from_candidate(self, candidate_type, value):
        Partner = self.env['res.partner'].sudo().with_context(active_test=False)
        raw_value = str(value).strip()
        if not raw_value:
            return False

        if candidate_type == 'partner_id':
            try:
                partner = Partner.browse(int(float(raw_value))).exists()
            except (TypeError, ValueError):
                partner = False
            return partner

        if candidate_type == 'customer_id':
            return self._find_partner_by_customer_id(raw_value)

        if candidate_type == 'xml_id':
            partner = self.env.ref(raw_value, raise_if_not_found=False)
            if partner and partner._name == 'res.partner':
                return partner.sudo()
            matches = self.env['ir.model.data'].sudo().search(
                [('model', '=', 'res.partner'), ('name', '=', raw_value.split('.')[-1])],
                limit=2,
            )
            if len(matches) == 1:
                return self.env[matches.model].browse(matches.res_id).sudo()
            return False

        if candidate_type == 'ref':
            matches = Partner.search([('ref', '=', raw_value)], limit=2)
            return self._single_record_or_false(matches, raw_value, 'partner.ref')

        if candidate_type == 'email':
            matches = Partner.search([('email', '=ilike', raw_value)], limit=2)
            return self._single_record_or_false(matches, raw_value, 'partner.email')

        if candidate_type in ('mobile', 'phone'):
            return self._find_partner_by_phone(raw_value)

        if candidate_type == 'vat':
            matches = Partner.search([('vat', '=ilike', raw_value)], limit=2)
            return self._single_record_or_false(matches, raw_value, 'partner.vat')

        if candidate_type == 'name':
            matches = Partner.search([('name', '=ilike', raw_value)], limit=2)
            return self._single_record_or_false(matches, raw_value, 'partner.name')

        return False

    def _find_partner_by_phone(self, raw_value):
        digits = self._normalize_phone(raw_value)
        if not digits:
            return False
        Partner = self.env['res.partner'].sudo().with_context(active_test=False)
        suffix = digits[-8:]
        matches = Partner.search(
            ['|', ('phone', 'ilike', suffix), ('mobile', 'ilike', suffix)],
            limit=20,
        )
        exact = matches.filtered(
            lambda partner: self._normalize_phone(partner.phone).endswith(digits)
            or self._normalize_phone(partner.mobile).endswith(digits)
        )
        return self._single_record_or_false(exact[:2], raw_value, 'partner.phone')

    def _find_partner_by_customer_id(self, raw_value):
        value = str(raw_value or '').strip()
        if not value:
            return False
        Partner = self.env['res.partner'].sudo().with_context(active_test=False)
        field_names = self._get_customer_id_partner_field_names()
        matches = Partner.browse()
        for field_name in field_names:
            field = Partner._fields.get(field_name)
            if not field or field.type not in ('char', 'text', 'integer', 'float'):
                continue
            domain_value = value
            if field.type in ('integer', 'float'):
                try:
                    domain_value = int(float(value))
                except (TypeError, ValueError):
                    continue
                operator = '='
            else:
                operator = '=ilike'
            current = Partner.search([(field_name, operator, domain_value)], limit=2)
            if current:
                if len(current) > 1:
                    raise UserError(
                        _('La búsqueda por ID Cliente "%(value)s" devolvió varios partners en el campo %(field)s.') % {
                            'value': value,
                            'field': field_name,
                        }
                    )
                if matches and matches[:1].id != current[:1].id:
                    raise UserError(
                        _('El ID Cliente "%(value)s" coincide con distintos partners según el campo configurado.') % {
                            'value': value,
                        }
                    )
                matches = current
        return matches[:1] if matches else False

    def _resolve_subscription_product(self, raw_value, cache, row_number):
        key = self._cacheable_value(raw_value)
        if key in cache:
            return cache[key]
        if self._is_empty_cell(raw_value):
            raise UserError(_('Fila %s: la columna plan es obligatoria.') % row_number)

        Product = self.env['product.product'].sudo().with_context(active_test=False)
        value = str(raw_value).strip()
        product = False
        try:
            product_id = int(float(value))
        except (TypeError, ValueError):
            product_id = 0
        if product_id > 0:
            product = Product.browse(product_id).exists()
            if product and not product.product_tmpl_id.recurring_invoice:
                raise UserError(_('Fila %s: el producto %s no es recurrente.') % (row_number, product.display_name))
            if product:
                cache[key] = product
                return product

        candidates = Product.search(
            [
                ('product_tmpl_id.recurring_invoice', '=', True),
                '|',
                '|',
                ('default_code', '=ilike', value),
                ('barcode', '=', value),
                ('name', '=ilike', value),
            ],
            limit=20,
        )
        resolved = self._pick_best_product_candidate(candidates, value)
        if resolved:
            cache[key] = resolved
            return resolved
        if candidates:
            raise UserError(
                _('Fila %(row)s: el plan "%(plan)s" coincide con varios productos recurrentes activos o válidos para la compañía.') % {
                    'row': row_number,
                    'plan': value,
                }
            )
        raise UserError(_('Fila %(row)s: no encontré un producto recurrente para "%(plan)s".') % {
            'row': row_number,
            'plan': value,
        })

    def _pick_best_product_candidate(self, candidates, raw_value):
        if not candidates:
            return False
        normalized_value = self._normalize_token(raw_value)
        exact = candidates.filtered(
            lambda product: self._normalize_token(product.display_name) == normalized_value
            or self._normalize_token(product.name) == normalized_value
            or self._normalize_token(product.default_code) == normalized_value
        )
        pools = [
            exact.filtered(lambda product: product.active),
            self._filter_products_for_company(exact.filtered(lambda product: product.active)),
            self._filter_products_for_company(exact),
            exact,
            self._filter_products_for_company(candidates.filtered(lambda product: product.active)),
            candidates.filtered(lambda product: product.active),
            self._filter_products_for_company(candidates),
            candidates,
        ]
        for pool in pools:
            resolved = self._resolve_unique_product_candidate(pool)
            if resolved:
                return resolved
        return False

    def _resolve_unique_product_candidate(self, products):
        products = products.exists()
        if not products:
            return False
        unique_ids = list(dict.fromkeys(products.ids))
        if len(unique_ids) == 1:
            return products.browse(unique_ids[0])

        template_ids = list(dict.fromkeys(products.mapped('product_tmpl_id').ids))
        if len(template_ids) == 1:
            active_same_template = products.filtered(lambda product: product.active)
            if active_same_template:
                return active_same_template.sorted(key=lambda product: product.id)[:1]
            return products.sorted(key=lambda product: product.id)[:1]
        return False

    def _filter_products_for_company(self, products):
        if not products:
            return products
        company_id = self.company_id.id
        filtered = products.filtered(
            lambda product: self._get_product_company_id(product) in (False, company_id)
        )
        return filtered or products

    def _get_product_company_id(self, product):
        if 'company_id' in product._fields and product.company_id:
            return product.company_id.id
        product_tmpl = product.product_tmpl_id if product else False
        if product_tmpl and 'company_id' in product_tmpl._fields and product_tmpl.company_id:
            return product_tmpl.company_id.id
        return False

    def _resolve_subscription_plan(self, raw_value, cache, row_number):
        if self._is_empty_cell(raw_value):
            return False
        key = self._cacheable_value(raw_value)
        if key in cache:
            return cache[key]

        value = str(raw_value).strip()
        candidate_model_names = self._get_subscription_plan_model_names()
        if not candidate_model_names:
            raise UserError(
                _('Fila %(row)s: no encontré ningún modelo de plan recurrente disponible en este entorno.') % {
                    'row': row_number,
                }
            )

        try:
            plan_id = int(float(value))
        except (TypeError, ValueError):
            plan_id = 0
        if plan_id > 0:
            for model_name in candidate_model_names:
                if model_name not in self.env.registry:
                    continue
                plan = self.env[model_name].sudo().browse(plan_id).exists()
                if plan:
                    cache[key] = plan
                    return plan

        normalized_value = self._normalize_token(value)
        record_matches = []
        seen_keys = set()
        for model_name in candidate_model_names:
            if model_name not in self.env.registry:
                continue
            Model = self.env[model_name].sudo().with_context(active_test=False)
            if 'name' in Model._fields:
                current = Model.search([('name', '=ilike', value)], limit=5)
            else:
                current = Model.browse()
            if not current and 'code' in Model._fields:
                current = Model.search([('code', '=ilike', value)], limit=5)
            current = current.filtered(
                lambda record: self._normalize_token(getattr(record, 'display_name', False) or getattr(record, 'name', False) or '') == normalized_value
                or self._normalize_token(getattr(record, 'name', False) or '') == normalized_value
            ) or current
            for record in current:
                record_key = (record._name, record.id)
                if record_key in seen_keys:
                    continue
                seen_keys.add(record_key)
                record_matches.append(record)

        if len(record_matches) == 1:
            cache[key] = record_matches[0]
            return record_matches[0]
        if len(record_matches) > 1:
            raise UserError(
                _('Fila %(row)s: el plan recurrente "%(plan)s" es ambiguo. Usa el ID exacto en subscription_plan.') % {
                    'row': row_number,
                    'plan': value,
                }
            )
        raise UserError(
            _('Fila %(row)s: no encontré el plan recurrente "%(plan)s".') % {
                'row': row_number,
                'plan': value,
            }
        )

    def _get_subscription_plan_model_names(self):
        model_names = set()
        checker = self._is_plan_model_name
        pos_order_model = self.env['pos.order'].sudo() if 'pos.order' in self.env.registry else False
        if pos_order_model:
            pos_checker = getattr(pos_order_model, '_wgs_is_plan_model_name', None)
            if callable(pos_checker):
                checker = pos_checker

        for model_name in ('sale.order', 'sale.order.line', 'product.product', 'product.template'):
            if model_name not in self.env.registry:
                continue
            for field in self.env[model_name]._fields.values():
                if field.type != 'many2one':
                    continue
                comodel_name = getattr(field, 'comodel_name', '')
                if checker(comodel_name):
                    model_names.add(comodel_name)
        return sorted(model_names)

    def _resolve_participants(self, raw_value, owner, row_number):
        if self._is_empty_cell(raw_value):
            return [owner.id] if owner else []
        raw_names = self._split_participants(raw_value)
        participant_ids = [owner.id] if owner else []
        seen_ids = set(participant_ids)
        for participant_name in raw_names:
            partner = self._find_partner_by_participant_name(participant_name)
            if owner and partner.id == owner.id:
                continue
            if partner.id in seen_ids:
                continue
            participant_ids.append(partner.id)
            seen_ids.add(partner.id)
        if not participant_ids:
            raise UserError(_('Fila %s: no se pudo resolver ningún participante.') % row_number)
        return participant_ids

    def _find_partner_by_participant_name(self, raw_value):
        value = str(raw_value or '').strip()
        if not value:
            return False
        Partner = self.env['res.partner'].sudo().with_context(active_test=False)
        normalized_value = self._normalize_token(value)
        matches = Partner.search([('name', '=ilike', value)], limit=10)
        exact = matches.filtered(
            lambda partner: self._normalize_token(partner.name) == normalized_value
            or self._normalize_token(partner.display_name) == normalized_value
        )
        if len(exact) == 1:
            return exact[:1]
        if len(exact) > 1 or len(matches) > 1:
            raise UserError(_('La búsqueda del participante "%s" devolvió varios partners.') % value)
        if matches:
            return matches[:1]

        token_candidates = self._search_participant_candidates(value)
        fuzzy_match = self._pick_best_participant_candidate(token_candidates, value)
        if fuzzy_match:
            return fuzzy_match
        raise UserError(_('No encontré al participante "%s".') % value)

    def _search_participant_candidates(self, raw_value):
        Partner = self.env['res.partner'].sudo().with_context(active_test=False)
        tokens = self._tokenize_name_for_match(raw_value)
        if not tokens:
            return Partner.browse()

        probe_tokens = []
        if tokens:
            probe_tokens.append(tokens[0])
        if len(tokens) > 1:
            probe_tokens.append(tokens[-1])
        if len(tokens) > 2:
            probe_tokens.append(tokens[-2])

        domain = []
        for token in probe_tokens:
            domain.extend(['|', ('name', 'ilike', token), ('display_name', 'ilike', token)])
        if domain and domain[0] == '|':
            domain = domain[1:]
        if not domain:
            return Partner.browse()

        candidates = Partner.search(domain, limit=80)
        if len(tokens) >= 2:
            filtered = candidates.filtered(
                lambda partner: all(
                    token in self._normalize_token(partner.name) or token in self._normalize_token(partner.display_name)
                    for token in tokens[:1] + tokens[-1:]
                )
            )
            if filtered:
                candidates = filtered
        return candidates

    def _pick_best_participant_candidate(self, candidates, raw_value):
        candidates = candidates.exists()
        if not candidates:
            return False

        scored = []
        for partner in candidates:
            score = self._score_participant_candidate(partner, raw_value)
            if score > 0:
                scored.append((score, partner))
        if not scored:
            return False

        scored.sort(key=lambda row: (-row[0], row[1].id))
        best_score, best_partner = scored[0]
        second_score = scored[1][0] if len(scored) > 1 else 0.0

        if best_score >= 0.93:
            return best_partner
        if best_score >= 0.86 and (best_score - second_score) >= 0.03:
            return best_partner
        if best_score >= 0.80 and second_score <= 0.72:
            return best_partner
        return False

    def _score_participant_candidate(self, partner, raw_value):
        input_normalized = self._normalize_token(raw_value)
        partner_normalized = self._normalize_token(partner.name or partner.display_name or '')
        display_normalized = self._normalize_token(partner.display_name or partner.name or '')
        if not input_normalized or not partner_normalized:
            return 0.0

        name_ratio = difflib.SequenceMatcher(None, input_normalized, partner_normalized).ratio()
        display_ratio = difflib.SequenceMatcher(None, input_normalized, display_normalized).ratio()

        input_token_list = self._tokenize_name_for_match(raw_value)
        partner_token_list = self._tokenize_name_for_match(partner.name or partner.display_name or '')
        input_tokens = set(input_token_list)
        partner_tokens = set(partner_token_list)
        token_overlap = 0.0
        if input_tokens and partner_tokens:
            token_overlap = len(input_tokens & partner_tokens) / float(max(len(input_tokens), len(partner_tokens)))

        starts_same = 1.0 if input_token_list and partner_token_list and input_token_list[0] == partner_token_list[0] else 0.0
        return max(name_ratio, display_ratio) * 0.75 + token_overlap * 0.2 + starts_same * 0.05

    def _resolve_recurring_context(self, product, subscription_plan=False, start_date=False, fallback_price=False):
        choice = {
            'price': float(product.list_price or 0.0),
            'plan': subscription_plan or False,
            'pricing_id': False,
            'next_billing_date': False,
        }
        pos_order_model = self.env['pos.order'].sudo() if 'pos.order' in self.env.registry else False
        if not pos_order_model:
            return choice

        try:
            fallback_amount = (
                self._parse_price_value(fallback_price, fallback=float(product.list_price or 0.0))
                if not self._is_empty_cell(fallback_price)
                else float(product.list_price or 0.0)
            )
            preferred_plan_id = subscription_plan.id if subscription_plan else False
            pricing_choice = pos_order_model._wgs_get_recurring_pricing_choice(
                product,
                fallback=fallback_amount,
                preferred_plan_id=preferred_plan_id,
                preferred_pricing_id=False,
            )
            plan_id = pricing_choice.get('plan_id') or False
            pricing_id = pricing_choice.get('pricing_id') or False
            resolved_plan = subscription_plan
            if not resolved_plan and plan_id:
                resolved_plan = pos_order_model._wgs_resolve_plan_record(
                    product=product,
                    plan_id=plan_id,
                    pricing_id=pricing_id,
                )
            next_billing_date = False
            if resolved_plan and start_date:
                next_billing_date = pos_order_model._wgs_get_plan_min_end_threshold(resolved_plan, start_date)
            return {
                'price': float(pricing_choice.get('price') or fallback_amount or 0.0),
                'plan': resolved_plan or False,
                'pricing_id': pricing_id or False,
                'next_billing_date': next_billing_date or False,
            }
        except Exception as error:  # pragma: no cover - defensive integration fallback
            _logger.warning(
                'WGS import: could not resolve recurring context for product=%s error=%s',
                product.id,
                error,
            )
            return choice

    def _create_subscription_order(
        self,
        partner,
        product,
        subscription_plan,
        start_date,
        end_date,
        state_value,
        price_unit,
        quantity,
        participant_ids,
        recurring_pricing_id,
        next_billing_date,
        source_key,
    ):
        order_model = self.env['sale.order'].sudo().with_company(self.company_id)
        line_values = self._build_sale_order_line_values(
            product,
            subscription_plan,
            price_unit,
            quantity,
            recurring_pricing_id=recurring_pricing_id,
        )
        order_values = self._build_sale_order_values(
            partner=partner,
            product=product,
            subscription_plan=subscription_plan,
            start_date=start_date,
            end_date=end_date,
            state_value=state_value,
            source_key=source_key,
            line_values=line_values,
        )
        order = order_model.create(order_values)
        if order.state in ('draft', 'sent'):
            order.action_confirm()
        self._write_subscription_metadata(
            order=order,
            partner=partner,
            product=product,
            subscription_plan=subscription_plan,
            start_date=start_date,
            end_date=end_date,
            state_value=state_value,
            price_unit=price_unit,
            quantity=quantity,
            participant_ids=participant_ids,
            recurring_pricing_id=recurring_pricing_id,
            next_billing_date=next_billing_date,
            source_key=source_key,
            allow_line_update=False,
        )
        return order

    def _update_existing_subscription_order(
        self,
        order,
        partner,
        product,
        subscription_plan,
        start_date,
        end_date,
        state_value,
        price_unit,
        quantity,
        participant_ids,
        recurring_pricing_id,
        next_billing_date,
        source_key,
    ):
        if order.state == 'cancel':
            raise UserError(
                _('La orden %s ya está cancelada; crea una nueva o reactívala manualmente antes de importar.') % order.display_name
            )
        self._write_subscription_metadata(
            order=order,
            partner=partner,
            product=product,
            subscription_plan=subscription_plan,
            start_date=start_date,
            end_date=end_date,
            state_value=state_value,
            price_unit=price_unit,
            quantity=quantity,
            participant_ids=participant_ids,
            recurring_pricing_id=recurring_pricing_id,
            next_billing_date=next_billing_date,
            source_key=source_key,
            allow_line_update=True,
        )
        if order.state in ('draft', 'sent'):
            order.action_confirm()
        return order

    def _write_subscription_metadata(
        self,
        order,
        partner,
        product,
        subscription_plan,
        start_date,
        end_date,
        state_value,
        price_unit,
        quantity,
        participant_ids,
        recurring_pricing_id,
        next_billing_date,
        source_key,
        allow_line_update,
    ):
        order = order.sudo().with_company(self.company_id)
        recurring_lines = order.order_line.filtered(
            lambda line: line.product_id and line.product_id.product_tmpl_id.recurring_invoice
        )
        if allow_line_update:
            if recurring_lines and recurring_lines[:1].product_id != product:
                raise UserError(
                    _('La orden %s ya existe pero usa un producto recurrente distinto: %s.') % (
                        order.display_name,
                        recurring_lines[:1].product_id.display_name,
                    )
                )
            if not recurring_lines:
                raise UserError(
                    _('La orden %s ya existe pero no tiene una línea recurrente para actualizar.') % order.display_name
                )

        write_values = {
            'partner_id': partner.id,
            'company_id': self.company_id.id,
            'wgs_import_source_key': source_key,
            'wgs_import_batch_name': self.batch_name,
        }
        if 'pricelist_id' in order._fields and partner.property_product_pricelist:
            write_values['pricelist_id'] = partner.property_product_pricelist.id
        if 'subscription_state' in order._fields:
            write_values['subscription_state'] = state_value
        if 'wgs_effective_start_date' in order._fields:
            write_values['wgs_effective_start_date'] = start_date
        self._assign_date_field(
            values=write_values,
            fields_map=order._fields,
            value_date=start_date,
            preferred_names=('start_date', 'date_start', 'subscription_start_date'),
        )
        self._assign_date_field(
            values=write_values,
            fields_map=order._fields,
            value_date=end_date,
            preferred_names=('date_end', 'end_date', 'subscription_end_date', 'recurring_end_date'),
        )
        if 'date_order' in order._fields:
            write_values['date_order'] = self._convert_for_field(start_date, order._fields['date_order'])
        self._assign_date_field(
            values=write_values,
            fields_map=order._fields,
            value_date=start_date,
            preferred_names=('first_contract_date', 'contract_date', 'date_contract'),
        )
        if subscription_plan:
            for field_name in ('subscription_plan_id', 'plan_id', 'recurring_plan_id'):
                if field_name in order._fields and field_name not in write_values:
                    write_values[field_name] = subscription_plan.id
        self._assign_many2one_value(
            values=write_values,
            fields_map=order._fields,
            value_id=subscription_plan.id if subscription_plan else False,
            preferred_names=('plan_id', 'subscription_plan_id', 'recurring_plan_id'),
            comodel_checker=self._is_plan_model_name,
        )
        order.write(write_values)

        if allow_line_update:
            line = recurring_lines[:1]
            line_values = {'product_id': product.id}
            qty_field = self._get_line_qty_field_name(line._fields)
            if qty_field:
                line_values[qty_field] = quantity
            if 'price_unit' in line._fields:
                line_values['price_unit'] = price_unit
            if subscription_plan:
                for field_name in ('subscription_plan_id', 'plan_id', 'recurring_plan_id'):
                    if field_name in line._fields and field_name not in line_values:
                        line_values[field_name] = subscription_plan.id
            self._assign_many2one_value(
                values=line_values,
                fields_map=line._fields,
                value_id=subscription_plan.id if subscription_plan else False,
                preferred_names=('subscription_plan_id', 'plan_id', 'recurring_plan_id'),
                comodel_checker=self._is_plan_model_name,
            )
            if recurring_pricing_id:
                for field_name in ('subscription_pricing_id', 'pricing_id', 'recurring_pricing_id'):
                    if field_name in line._fields and field_name not in line_values:
                        line_values[field_name] = recurring_pricing_id
            self._assign_many2one_value(
                values=line_values,
                fields_map=line._fields,
                value_id=recurring_pricing_id,
                preferred_names=('subscription_pricing_id', 'pricing_id', 'recurring_pricing_id'),
                comodel_checker=self._is_pricing_model_name,
            )
            line.write(line_values)

        self._sync_subscription_runtime_metadata(
            order=order,
            participant_ids=participant_ids,
            contract_date=start_date,
            subscription_start_date=start_date,
            subscription_end_date=end_date,
            next_billing_date=next_billing_date,
        )
        self._sync_order_participants(order, participant_ids, quantity, product)
        order._ensure_subscription_owner_is_participant()

    def _build_sale_order_values(
        self,
        partner,
        product,
        subscription_plan,
        start_date,
        end_date,
        state_value,
        source_key,
        line_values,
    ):
        order_model = self.env['sale.order'].sudo().with_company(self.company_id)
        values = {
            'partner_id': partner.id,
            'company_id': self.company_id.id,
            'wgs_import_source_key': source_key,
            'wgs_import_batch_name': self.batch_name,
            'order_line': [Command.create(line_values)],
        }
        if 'origin' in order_model._fields:
            values['origin'] = self.batch_name
        if 'client_order_ref' in order_model._fields:
            values['client_order_ref'] = self.batch_name
        if 'pricelist_id' in order_model._fields and partner.property_product_pricelist:
            values['pricelist_id'] = partner.property_product_pricelist.id
        if 'date_order' in order_model._fields:
            values['date_order'] = self._convert_for_field(start_date, order_model._fields['date_order'])
        self._assign_date_field(
            values=values,
            fields_map=order_model._fields,
            value_date=start_date,
            preferred_names=('first_contract_date', 'contract_date', 'date_contract'),
        )
        if 'subscription_state' in order_model._fields:
            values['subscription_state'] = state_value
        if 'wgs_effective_start_date' in order_model._fields:
            values['wgs_effective_start_date'] = start_date
        self._assign_date_field(
            values=values,
            fields_map=order_model._fields,
            value_date=start_date,
            preferred_names=('start_date', 'date_start', 'subscription_start_date'),
        )
        self._assign_date_field(
            values=values,
            fields_map=order_model._fields,
            value_date=end_date,
            preferred_names=('date_end', 'end_date', 'subscription_end_date', 'recurring_end_date'),
        )
        if subscription_plan:
            for field_name in ('subscription_plan_id', 'plan_id', 'recurring_plan_id'):
                if field_name in order_model._fields and field_name not in values:
                    values[field_name] = subscription_plan.id
        self._assign_many2one_value(
            values=values,
            fields_map=order_model._fields,
            value_id=subscription_plan.id if subscription_plan else False,
            preferred_names=('plan_id', 'subscription_plan_id', 'recurring_plan_id'),
            comodel_checker=self._is_plan_model_name,
        )
        return values

    def _sync_subscription_runtime_metadata(
        self,
        order,
        participant_ids,
        contract_date=False,
        subscription_start_date=False,
        subscription_end_date=False,
        next_billing_date=False,
    ):
        pos_order_model = self.env['pos.order'].sudo() if 'pos.order' in self.env.registry else False
        if not pos_order_model:
            return
        sync_method = getattr(pos_order_model, '_wgs_sync_subscription_metadata', None)
        if not callable(sync_method):
            return
        sync_method(
            sale_order=order,
            participant_ids=participant_ids or [],
            contract_date=contract_date,
            subscription_start_date=subscription_start_date,
            subscription_end_date=subscription_end_date,
            next_billing_date=next_billing_date,
        )

    def _build_sale_order_line_values(self, product, subscription_plan, price_unit, quantity, recurring_pricing_id=False):
        line_model = self.env['sale.order.line'].sudo()
        values = {'product_id': product.id}
        if 'name' in line_model._fields:
            values['name'] = product.display_name
        if 'product_uom' in line_model._fields and product.uom_id:
            values['product_uom'] = product.uom_id.id
        qty_field = self._get_line_qty_field_name(line_model._fields)
        if qty_field:
            values[qty_field] = quantity
        if 'price_unit' in line_model._fields:
            values['price_unit'] = price_unit
        if subscription_plan:
            for field_name in ('subscription_plan_id', 'plan_id', 'recurring_plan_id'):
                if field_name in line_model._fields and field_name not in values:
                    values[field_name] = subscription_plan.id
        self._assign_many2one_value(
            values=values,
            fields_map=line_model._fields,
            value_id=subscription_plan.id if subscription_plan else False,
            preferred_names=('subscription_plan_id', 'plan_id', 'recurring_plan_id'),
            comodel_checker=self._is_plan_model_name,
        )
        if recurring_pricing_id:
            for field_name in ('subscription_pricing_id', 'pricing_id', 'recurring_pricing_id'):
                if field_name in line_model._fields and field_name not in values:
                    values[field_name] = recurring_pricing_id
        self._assign_many2one_value(
            values=values,
            fields_map=line_model._fields,
            value_id=recurring_pricing_id,
            preferred_names=('subscription_pricing_id', 'pricing_id', 'recurring_pricing_id'),
            comodel_checker=self._is_pricing_model_name,
        )
        return values

    def _resolve_subscription_state_value(self, mode):
        field = self.env['sale.order']._fields.get('subscription_state')
        if not field:
            return False
        selection = field.selection
        if callable(selection):
            try:
                selection = selection(self.env['sale.order'])
            except TypeError:
                selection = selection(self.env)
        selection = selection or []
        if mode == 'closed':
            wanted = ('close', 'closed', 'cerrada', 'cerrado')
        elif mode == 'renew':
            wanted = ('renew', 'to renew', 'por renovar', 'progress', 'en progreso')
        else:
            wanted = ('progress', 'in progress', 'en progreso', 'renew', 'por renovar')
        for value, label in selection:
            haystack = ' '.join(filter(None, [str(value).lower(), str(label).lower()]))
            if any(token in haystack for token in wanted):
                return value
        return False

    def _resolve_import_subscription_state_value(
        self,
        start_date,
        end_date,
        today,
        active_state_value,
    ):
        return active_state_value

    def _ensure_subscription_plan_ready(self, product, subscription_plan, row_number):
        if subscription_plan:
            return
        raise UserError(
            _(
                'Fila %(row)s: no pude resolver el plan recurrente para "%(product)s". '
                'Agrega la columna subscription_plan en el Excel con el nombre o ID del plan recurrente.'
            )
            % {
                'row': row_number,
                'product': product.display_name,
            }
        )

    def _assign_date_field(self, values, fields_map, value_date, preferred_names):
        if not value_date:
            return
        for field_name in preferred_names:
            if field_name in fields_map and field_name not in values:
                values[field_name] = self._convert_for_field(value_date, fields_map[field_name])
                return

    def _assign_many2one_value(self, values, fields_map, value_id, preferred_names, comodel_checker=None):
        value_id = int(value_id or 0)
        if value_id <= 0:
            return
        for field_name in preferred_names:
            field = fields_map.get(field_name)
            if not field or field.type != 'many2one':
                continue
            values[field_name] = value_id
            return
        if not comodel_checker:
            return
        for field_name, field in fields_map.items():
            if field_name in values or field.type != 'many2one':
                continue
            normalized_name = (field_name or '').lower()
            heuristic_name_match = (
                ('plan' in normalized_name)
                or ('recurr' in normalized_name)
                or ('period' in normalized_name)
                or ('pricing' in normalized_name)
                or ('price' in normalized_name)
            )
            if comodel_checker(getattr(field, 'comodel_name', '')) or heuristic_name_match:
                values[field_name] = value_id
                return

    def _is_plan_model_name(self, model_name):
        model_name = (model_name or '').lower()
        if not model_name:
            return False
        if 'subscription.plan' in model_name or 'recurring.plan' in model_name:
            return True
        if 'recurr' in model_name:
            return True
        if 'subscription' in model_name and ('period' in model_name or 'template' in model_name):
            return True
        return model_name.endswith('.plan')

    def _is_pricing_model_name(self, model_name):
        model_name = (model_name or '').lower()
        if not model_name:
            return False
        if 'subscription.pricing' in model_name or 'recurring.pricing' in model_name:
            return True
        return 'subscription' in model_name and 'price' in model_name

    def _convert_for_field(self, value_date, field):
        if not value_date:
            return False
        if field.type == 'datetime':
            return datetime.combine(value_date, time.min)
        return value_date

    def _parse_date_value(self, raw_value, field_label, row_number):
        if self._is_empty_cell(raw_value):
            return False
        if isinstance(raw_value, datetime):
            return raw_value.date()
        if isinstance(raw_value, date):
            return raw_value
        text = str(raw_value).strip()
        if not text:
            return False
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%m/%d/%Y', '%m-%d-%Y'):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        raise UserError(
            _('Fila %(row)s: no pude interpretar la fecha de %(label)s "%(value)s".') % {
                'row': row_number,
                'label': field_label,
                'value': text,
            }
        )

    def _parse_quantity_value(self, raw_value):
        if self._is_empty_cell(raw_value):
            return 1.0
        try:
            quantity = float(raw_value)
        except (TypeError, ValueError):
            raise UserError(_('Cantidad inválida: %s') % raw_value)
        return quantity if quantity > 0 else 1.0

    def _parse_price_value(self, raw_value, fallback):
        if self._is_empty_cell(raw_value):
            return fallback
        if isinstance(raw_value, str):
            raw_value = raw_value.replace(',', '').strip()
        try:
            return float(raw_value)
        except (TypeError, ValueError):
            raise UserError(_('Precio inválido: %s') % raw_value)

    def _build_source_key(self, partner, product, start_date):
        return 'wgs-sub:%s:%s:%s:%s' % (
            self.company_id.id,
            partner.id,
            product.id,
            fields.Date.to_string(start_date),
        )

    def _sync_order_participants(self, order, participant_ids, quantity, product):
        if 'participant_ids' not in order._fields or not participant_ids:
            return
        max_total = int((float(quantity or 0.0) or 0.0) * int(product.product_tmpl_id.max_participants_total or 0))
        if max_total and len(participant_ids) > max_total:
            raise UserError(
                _(
                    'No puedes asignar %(current)s participantes para %(product)s. El máximo permitido es %(max)s.'
                )
                % {
                    'current': len(participant_ids),
                    'product': product.display_name,
                    'max': max_total,
                }
            )
        order.write({'participant_ids': [Command.set(participant_ids)]})

    def _normalize_phone(self, value):
        return ''.join(ch for ch in str(value or '') if ch.isdigit())

    def _split_participants(self, raw_value):
        if isinstance(raw_value, str):
            text = raw_value
        else:
            text = str(raw_value or '')
        text = text.replace(';', ',')
        return [part.strip() for part in text.split(',') if part and part.strip()]

    def _tokenize_name_for_match(self, value):
        normalized = self._normalize_token(value)
        return [token for token in normalized.split('_') if token and len(token) >= 2]

    def _get_customer_id_partner_field_names(self):
        return ['x_studio_id_de_cliente']

    def _cacheable_value(self, value):
        if isinstance(value, str):
            return self._normalize_token(value)
        if isinstance(value, (int, float)):
            return str(value)
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return fields.Date.to_string(value)
        return str(value or '')

    def _normalize_token(self, value):
        text = str(value or '').strip().lower()
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(ch for ch in text if not unicodedata.combining(ch))
        cleaned = []
        for char in text:
            if char.isalnum():
                cleaned.append(char)
            else:
                cleaned.append('_')
        return ''.join(cleaned).strip('_')

    def _is_empty_cell(self, value):
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        return False

    def _get_line_qty_field_name(self, fields_map):
        for field_name in ('product_uom_qty', 'quantity', 'qty'):
            if field_name in fields_map:
                return field_name
        return False

    def _single_record_or_false(self, records, value, label):
        if not records:
            return False
        if len(records) > 1:
            raise UserError(_('La búsqueda por %(label)s "%(value)s" devolvió varios registros.') % {
                'label': label,
                'value': value,
            })
        return records

    def default_get(self, fields_list):
        values = super().default_get(fields_list)
        if 'batch_name' in fields_list and not values.get('batch_name') and self._context.get('default_file_name'):
            values['batch_name'] = os.path.splitext(self._context['default_file_name'])[0]
        return values
